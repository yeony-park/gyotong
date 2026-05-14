"""
KTX 공실률 분석 - 데이터 전처리 스크립트
입력: 4_고속철도_여객_수송동향.xlsx
출력: ktx_wide.csv, ktx_long.csv
"""

import pandas as pd
import numpy as np
from pathlib import Path

# ── 경로 설정 ──────────────────────────────────────────
INPUT_PATH = "4_고속철도_여객_수송동향.xlsx"
OUTPUT_WIDE = "ktx_wide.csv"
OUTPUT_LONG = "ktx_long.csv"


def load_raw(path: str) -> pd.DataFrame:
    """엑셀 파일에서 Wide DataFrame 생성"""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # 날짜 추출 (행3, 컬럼2~)
    dates, date_cols = [], []
    for c in range(2, ws.max_column + 1):
        val = ws.cell(row=3, column=c).value
        if val and str(val).strip():
            dates.append(str(val).strip())
            date_cols.append(c)

    # 지표명 추출 (컬럼1, 행4~)
    indicators, indicator_rows = [], []
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip() not in ["", "출처:"]:
            indicators.append(str(val).strip().replace("\xa0", " ").strip())
            indicator_rows.append(r)

    # 데이터 추출
    data = {}
    for ind, r in zip(indicators, indicator_rows):
        row_data = []
        for c in date_cols:
            val = ws.cell(row=r, column=c).value
            if val == "-" or val is None:
                row_data.append(np.nan)
            else:
                try:
                    row_data.append(float(str(val).replace(",", "")))
                except Exception:
                    row_data.append(np.nan)
        data[ind] = row_data

    df = pd.DataFrame(data, index=dates)
    df.index = pd.to_datetime(
        df.index.str.replace("월", ""), format="%Y%m"
    )
    df.index.name = "yearmonth"
    return df


def add_vacancy_rate(df: pd.DataFrame) -> pd.DataFrame:
    """공실률 = 100 - 이용률 컬럼 추가"""
    mapping = {
        "KTX이용률":       "KTX_공실률",
        "경부선 KTX이용률": "경부선_KTX_공실률",
        "호남선 KTX이용률": "호남선_KTX_공실률",
        "경전선  KTX이용률":"경전선_KTX_공실률",
        "전라선 KTX이용률": "전라선_KTX_공실률",
        "동해선 KTX이용률": "동해선_KTX_공실률",
    }
    for src, dst in mapping.items():
        if src in df.columns:
            df[dst] = 100 - df[src]
    return df


def add_time_features(df: pd.DataFrame) -> pd.DataFrame:
    """날짜 파생변수 추가"""
    df["연도"] = df.index.year
    df["월"] = df.index.month
    df["분기"] = df.index.quarter
    df["계절"] = df["월"].map({
        12: "겨울", 1: "겨울", 2: "겨울",
        3: "봄",   4: "봄",   5: "봄",
        6: "여름", 7: "여름", 8: "여름",
        9: "가을", 10: "가을", 11: "가을",
    })
    # 이벤트 플래그
    df["코로나기간"] = ((df.index >= "2020-01") & (df.index <= "2022-06")).astype(int)
    df["SRT개통후"] = (df.index >= "2016-12").astype(int)
    return df


def to_long(df: pd.DataFrame) -> pd.DataFrame:
    """노선별 Long 형태 변환"""
    노선_map = {
        "경부선": ("경부선 KTX여객수", "경부선 KTX이용률", "경부선_KTX_공실률"),
        "호남선": ("호남선 KTX여객수", "호남선 KTX이용률", "호남선_KTX_공실률"),
        "경전선": ("경전선 KTX여객수", "경전선  KTX이용률", "경전선_KTX_공실률"),
        "전라선": ("전라선 KTX여객수", "전라선 KTX이용률", "전라선_KTX_공실률"),
        "동해선": ("동해선 KTX여객수", "동해선 KTX이용률", "동해선_KTX_공실률"),
    }
    rows = []
    for idx, row in df.iterrows():
        for 노선, (여객col, 이용률col, 공실률col) in 노선_map.items():
            공실률 = row.get(공실률col, np.nan)
            if pd.isna(공실률):
                continue

            rows.append({
                "yearmonth":      idx,
                "연도":           row["연도"],
                "월":             row["월"],
                "분기":           row["분기"],
                "계절":           row["계절"],
                "코로나기간":     row["코로나기간"],
                "SRT개통후":      row["SRT개통후"],
                "노선":           노선,
                "여객수_천명":    row.get(여객col, np.nan),
                "KTX이용률":      row.get(이용률col, np.nan),
                "공실률":         공실률,
                "초과수요":       1 if 공실률 < 0 else 0,
                "공실상태":       (
                    "초과수요"
                    if 공실률 < 0
                    else "여유없음"
                    if 공실률 < 10
                    else "공실있음"
                ),
                "KTX전체이용률":  row.get("KTX이용률", np.nan),
                "KTX전체공실률":  row.get("KTX_공실률", np.nan),
            })
    df_long = pd.DataFrame(rows)
    return df_long.reset_index(drop=True)


def main():
    print("1. 원본 데이터 로드...")
    df_wide = load_raw(INPUT_PATH)

    print("2. 공실률 컬럼 추가...")
    df_wide = add_vacancy_rate(df_wide)

    print("3. 시간 파생변수 추가...")
    df_wide = add_time_features(df_wide)

    print("4. Long 형태 변환...")
    df_long = to_long(df_wide)

    print("5. CSV 저장...")
    df_wide.to_csv(OUTPUT_WIDE, encoding="utf-8-sig")
    df_long.to_csv(OUTPUT_LONG, index=False, encoding="utf-8-sig")

    print(f"\n✅ 완료")
    print(f"   Wide: {df_wide.shape} → {OUTPUT_WIDE}")
    print(f"   Long: {df_long.shape} → {OUTPUT_LONG}")
    print(f"\n노선별 공실률 기초통계:")
    print(df_long.groupby("노선")["공실률"].describe().round(1).to_string())


if __name__ == "__main__":
    main()
